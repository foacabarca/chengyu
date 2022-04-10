#!python3
# coding=utf8

"""
Project:表格上的皇后
Author:42024130-宋加运
Date:2022.4.10
"""

from openpyxl import *
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

wb = load_workbook("E:/PythonProject/nqueen.xlsx")  # 加载excel文件
ws = wb["Sheet1"]
wb.remove(ws)  # 删除原有表单
sheet = wb.create_sheet("Sheet1")  # 创建新表单
n = 8  # 设置n皇后的皇后数目，可更改
queenpos = [0 for i in range(n)]  # 用于存放n个皇后的位置，第i+1个皇后在第queenpos[i]+1列
answer_num = 0  # n皇后解的个数
row_num = 0  # 行数
height = 35  # 设置单元格高度
width = 30  # 设置单元格宽度
# 设置字体居中
alignment = Alignment(horizontal="center", vertical="center", text_rotation=0, wrap_text=True)
# 设置字体格式
font1 = Font(name="微软雅黑", size=20, bold=True, italic=False, color="000000")
font2 = Font(name="微软雅黑", size=20, bold=True, italic=False, color="FFFFFF")


def fillcolor0():
    """
    把空白行填充为绿色，用于分隔n皇后的每一种答案
    """
    global row_num
    global sheet
    global height
    global width
    sheet.row_dimensions[row_num + 1].height = height
    sheet.row_dimensions[row_num + 1].width = width
    for i in range(n):
        sheet.cell(row=row_num + 1, column=i + 1).fill = \
            PatternFill(fgColor="20B2AA", fill_type="solid")


def fillcolor1():
    """
    奇数行：偶数列填充黑色，字体颜色为白色
    偶数行：奇数列填充黑色，字体颜色为白色
    其他单元格字体颜色为白色
    另外设置单元格的宽和高
    """
    global height, width, row_num, font1, font2, alignment, sheet
    sheet.row_dimensions[row_num + 1].height = height
    sheet.row_dimensions[row_num + 1].width = width
    for i in range(n):
        sheet.cell(row=row_num + 1, column=i + 1).alignment = alignment
        if row_num % 2 == (i + 1) % 2:
            sheet.cell(row=row_num + 1, column=i + 1).fill = \
                PatternFill(fgColor="000000", fill_type="solid")
            sheet.cell(row=row_num + 1, column=i + 1).font = font2
        else:
            sheet.cell(row=row_num + 1, column=i + 1).font = font1


def nqueen(k):
    """
    迭代法解决n皇后问题，
    每输出一行就设置该行的字体，格式，颜色等等
    """
    global row_num, answer_num, queenpos, n
    pos = [0 for l in range(n)]
    if k == n:
        for i in range(n):
            # 每生成一个位置，就新增一行，在特定位置输出皇后
            row = ["" for l in range(n)]
            row[queenpos[i]] = "♕"
            sheet.append(row)
            fillcolor1()
            row_num = row_num + 1
        # 新增空白内容行，分隔答案
        sheet.append(["" for l in range(n)])
        fillcolor0()
        answer_num = answer_num + 1
        row_num = row_num + 1
        print(queenpos)
        return
    for i in range(n):
        j = 0
        while j < k:
            # 判断是否能放置皇后
            if queenpos[j] == i or abs(queenpos[j] - i) == abs(k - j):
                break
            j = j + 1
        if j == k:
            queenpos[k] = i
            nqueen(k + 1)


nqueen(0)
print("答案个数为", answer_num)  # 控制台输出n皇后答案的个数
wb.save("E:/PythonProject/nqueen.xlsx")  # 保存
